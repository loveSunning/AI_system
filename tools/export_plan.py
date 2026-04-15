#!/usr/bin/env python3
"""Export the weekly AI-system Excel plan into a lightweight Markdown table."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def collect_rows(workbook_path: Path) -> list[tuple[str, str, str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: list[tuple[str, str, str, str]] = []
    current_module = ""
    current_topic = ""

    for row in sheet.iter_rows(min_row=5, values_only=True):
        module = row[0] or current_module
        topic = row[1] or current_topic
        week = row[2]
        focus = row[4]
        milestone = row[7]

        if not week:
            continue

        current_module = module
        current_topic = topic
        rows.append((str(week), str(module), str(topic), f"{focus} | {milestone}"))

    return rows


def render_markdown(rows: list[tuple[str, str, str, str]]) -> str:
    lines = [
        "| 周次 | 模块 | 主题 | 学习重点 / 验收 |",
        "| --- | --- | --- | --- |",
    ]
    for week, module, topic, detail in rows:
        safe_detail = detail.replace("\n", "<br>")
        lines.append(f"| {week} | {module} | {topic} | {safe_detail} |")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Path to AI-system-plan.xlsx")
    args = parser.parse_args()

    rows = collect_rows(args.input)
    print(render_markdown(rows))


if __name__ == "__main__":
    main()
